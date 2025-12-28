import os
from openpyxl import Workbook, load_workbook

DATA_DIR = "data"
DATA_PATH = os.path.join(DATA_DIR, "prizes.xlsx")

def init_prize_file():
    os.makedirs(DATA_DIR, exist_ok=True)

    if not os.path.exists(DATA_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["경품명", "수량"])
        wb.save(DATA_PATH)

def get_prizes():
    init_prize_file()
    wb = load_workbook(DATA_PATH)
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))

def decrease_prize(prize_name):
    wb = load_workbook(DATA_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == prize_name:
            if row[1].value <= 0:
                return False
            row[1].value -= 1
            wb.save(DATA_PATH)
            return True
    return False
