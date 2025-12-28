import random
from openpyxl import load_workbook

from services.participant_service import get_draw_candidates, set_winner
from services.prize_service import decrease_prize

DATA_PATH = "data/participants.xlsx"

def draw_with_prize(prize_name):
    candidates = get_draw_candidates()

    if not candidates:
        return None

    selected_row = random.choice(candidates)

    # 경품 수량 차감
    if not decrease_prize(prize_name):
        return "NO_PRIZE"

    wb = load_workbook(DATA_PATH)
    ws = wb.active
    name = ws.cell(row=selected_row, column=1).value
    org = ws.cell(row=selected_row, column=2).value

    set_winner(selected_row, prize_name)

    return {
        "name": name,
        "organization": org,
        "prize": prize_name
    }

def get_candidate_names():
    from openpyxl import load_workbook
    wb = load_workbook("data/participants.xlsx")
    ws = wb.active

    return [
        row[0]
        for row in ws.iter_rows(min_row=2, values_only=True)
        if not row[2]
    ]

