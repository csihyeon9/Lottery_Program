import os
from openpyxl import Workbook, load_workbook

DATA_DIR = "data"
DATA_PATH = os.path.join(DATA_DIR, "participants.xlsx")

def init_participant_file():
    os.makedirs(DATA_DIR, exist_ok=True)

    if not os.path.exists(DATA_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["이름", "소속", "당첨경품"])
        wb.save(DATA_PATH)

def add_participant(name, organization):
    init_participant_file()
    wb = load_workbook(DATA_PATH)
    ws = wb.active
    ws.append([name, organization, ""])
    wb.save(DATA_PATH)

def get_draw_candidates():
    init_participant_file()
    wb = load_workbook(DATA_PATH)
    ws = wb.active

    candidates = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[2]:
            candidates.append(idx)
    return candidates

def set_winner(row_index, prize_name):
    wb = load_workbook(DATA_PATH)
    ws = wb.active
    ws.cell(row=row_index, column=3).value = prize_name
    wb.save(DATA_PATH)
